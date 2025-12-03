#!/usr/bin/env python3
"""
Convert Neal's CNC Excel export to Django fixture JSON.

This script parses the Excel file, filters data to keep 100 Projects and related records,
handles contact name mismatches interactively, and generates a JSON fixture file for
import via Django's loaddata command.

Usage:
    python convert_neals_data.py nealsdata/company-export.xlsx
    python convert_neals_data.py nealsdata/company-export.xlsx --output my_data.json
    python convert_neals_data.py nealsdata/company-export.xlsx --non-interactive
    python convert_neals_data.py nealsdata/company-export.xlsx --dry-run
"""

import argparse
import json
import sys
from datetime import datetime, timedelta
from decimal import Decimal
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


class ContactMismatchHandler:
    """Handles interactive contact name mismatch resolution."""

    def __init__(self, interactive: bool = True):
        self.interactive = interactive
        self.decisions = {}  # Cache decisions for same mismatch

    def prompt_for_decision(
        self,
        business: str,
        existing_name: str,
        existing_email: str,
        referenced_name: str,
        sheet: str,
        row: int,
        context: str
    ) -> str:
        """
        Prompt user for decision on contact mismatch.
        Returns: 'update', 'create', or 'map'
        """
        key = (business, existing_name, referenced_name)

        # Return cached decision if we've seen this before
        if key in self.decisions:
            return self.decisions[key]

        # Non-interactive mode: always map to existing
        if not self.interactive:
            self.decisions[key] = 'map'
            return 'map'

        # Interactive prompt
        print("\n" + "=" * 64)
        print("Contact Mismatch Detected")
        print("=" * 64)
        print(f"Business:        {business}")
        print(f"Contacts Sheet:  {existing_name} ({existing_email})")
        print()
        print(f'Referenced as:   "{referenced_name}"')
        print(f"Found in:        {sheet} sheet, Row {row}")
        print(f"Context:         {context}")
        print()
        print("How should this be handled?")
        print(f'  [1] Update - Change contact name to "{referenced_name}"')
        print(f'  [2] Create - Create new contact "{referenced_name}" for this business')
        print(f'  [3] Map - Use existing contact "{existing_name}" as-is')
        print()

        while True:
            choice = input("Your choice (1/2/3): ").strip()
            if choice == '1':
                decision = 'update'
                break
            elif choice == '2':
                decision = 'create'
                break
            elif choice == '3':
                decision = 'map'
                break
            else:
                print("Invalid choice. Please enter 1, 2, or 3.")

        print("=" * 64)

        self.decisions[key] = decision
        return decision


class ExcelDataLoader:
    """Loads and parses Excel sheets into structured data."""

    def __init__(self, excel_path: str, verbose: bool = False):
        self.excel_path = excel_path
        self.verbose = verbose
        self.wb = None
        self.sheets_data = {}

    def load(self):
        """Load all required sheets into memory."""
        if self.verbose:
            print(f"Loading Excel file: {self.excel_path}")

        self.wb = openpyxl.load_workbook(self.excel_path, data_only=True)

        sheets_to_load = [
            'Contacts', 'Projects', 'Invoices', 'Estimates',
            'Bills', 'Tasks', 'Timeslips', 'Price List Items'
        ]

        for sheet_name in sheets_to_load:
            if sheet_name in self.wb.sheetnames:
                self.sheets_data[sheet_name] = self._load_sheet(sheet_name)
                if self.verbose:
                    print(f"  Loaded {sheet_name}: {len(self.sheets_data[sheet_name])} rows")
            else:
                print(f"Warning: Sheet '{sheet_name}' not found in workbook")
                self.sheets_data[sheet_name] = []

        self.wb.close()

    def _load_sheet(self, sheet_name: str) -> List[Dict]:
        """Load sheet into list of dictionaries."""
        ws = self.wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            return []

        headers = rows[0]
        data = []

        for row_idx, row_values in enumerate(rows[1:], start=2):
            row_dict = {
                '_row': row_idx,
                '_sheet': sheet_name
            }
            for idx, header in enumerate(headers):
                if header and idx < len(row_values):
                    row_dict[header] = row_values[idx]
            data.append(row_dict)

        return data


class NealsDataConverter:
    """Main converter class that orchestrates the conversion process."""

    def __init__(
        self,
        excel_path: str,
        output_path: str = "neals_data.json",
        base_fixture_path: str = "fixtures/job_data/01_base.json",
        interactive: bool = True,
        dry_run: bool = False,
        verbose: bool = False
    ):
        self.excel_path = excel_path
        self.output_path = output_path
        self.base_fixture_path = base_fixture_path
        self.interactive = interactive
        self.dry_run = dry_run
        self.verbose = verbose

        self.loader = ExcelDataLoader(excel_path, verbose)
        self.contact_handler = ContactMismatchHandler(interactive)

        # Data structures
        self.fixture_data = []
        self.pk_counters = {}  # Track next PK for each model

        # Load base fixtures first
        self._load_base_fixtures()

        # Lookup mappings
        self.business_map = {}  # org_name -> business_pk
        self.contact_map = {}  # (org_name, contact_name) -> contact_pk
        self.job_map = {}  # project_name -> (job_pk, workorder_pk or None)
        self.task_map = {}  # (project_name, task_name) -> task_pk

        # Filtered data
        self.filtered_contacts = []
        self.filtered_projects = []
        self.filtered_invoices = []
        self.filtered_estimates = []
        self.filtered_bills = []
        self.filtered_tasks = []
        self.filtered_timeslips = []
        self.filtered_price_list = []

        # Contact updates/creations
        self.contact_updates = {}  # pk -> new_name
        self.new_contacts = []  # List of new contact dicts to create

    def get_next_pk(self, model: str) -> int:
        """Get next primary key for a model type."""
        if model not in self.pk_counters:
            self.pk_counters[model] = 1
        pk = self.pk_counters[model]
        self.pk_counters[model] += 1
        return pk

    def add_fixture(self, model: str, pk: int, fields: Dict):
        """Add an object to the fixture data."""
        self.fixture_data.append({
            "model": model,
            "pk": pk,
            "fields": fields
        })

    def _load_base_fixtures(self):
        """Load existing base fixtures and initialize PK counters."""
        import os

        if not os.path.exists(self.base_fixture_path):
            if self.verbose:
                print(f"Base fixture file not found: {self.base_fixture_path}")
                print("Starting with empty fixture set")
            return

        if self.verbose:
            print(f"Loading base fixtures from: {self.base_fixture_path}")

        with open(self.base_fixture_path, 'r') as f:
            base_fixtures = json.load(f)

        # Add base fixtures to our fixture data
        self.fixture_data.extend(base_fixtures)

        # Calculate max PK for each model type to initialize counters
        max_pks = {}
        for fixture in base_fixtures:
            model = fixture['model']
            pk = fixture['pk']

            # Only track numeric PKs (some models like Configuration use string PKs)
            if isinstance(pk, int):
                if model not in max_pks:
                    max_pks[model] = pk
                else:
                    max_pks[model] = max(max_pks[model], pk)

        # Initialize PK counters to start after existing data
        for model, max_pk in max_pks.items():
            self.pk_counters[model] = max_pk + 1

        if self.verbose:
            print(f"  Loaded {len(base_fixtures)} base fixtures")
            print(f"  Initialized PK counters for {len(max_pks)} models:")
            for model, next_pk in sorted(self.pk_counters.items()):
                print(f"    {model}: starting at PK {next_pk}")
            print()

    def convert(self):
        """Main conversion process."""
        print("=" * 70)
        print("Neal's CNC Data Converter")
        print("=" * 70)
        print()

        # Phase 1: Load data
        print("[Phase 1] Loading Excel data...")
        self.loader.load()
        print()

        # Phase 2: Filter data
        print("[Phase 2] Filtering data...")
        self._filter_data()
        print()

        # Phase 3: Build objects (includes interactive contact resolution)
        print("[Phase 3] Building Django objects...")
        self._build_all_objects()
        print()

        # Summary
        self._print_summary()

        # Phase 4: Write JSON
        if not self.dry_run:
            self._write_json()
            print()
            print("=" * 70)
            print(f"✓ JSON fixture written to: {self.output_path}")
            print()
            print("To import into Django, run:")
            print(f"  python manage.py loaddata {self.output_path}")
            print("=" * 70)
        else:
            print()
            print("=" * 70)
            print("DRY RUN - No file written")
            print("=" * 70)

    def _filter_data(self):
        """Filter data to keep only relevant records."""
        # Keep all 100 Projects
        self.filtered_projects = self.loader.sheets_data.get('Projects', [])

        # Build set of project names for filtering
        project_names = {p.get('Name') for p in self.filtered_projects if p.get('Name')}

        if self.verbose:
            print(f"  Keeping {len(self.filtered_projects)} projects")
            print(f"  Project names set: {len(project_names)} unique names")

        # Filter other data based on project references
        # or has a recent date

        cutoff_date = datetime(2025, 10, 1)

        # Bills, Invoices and Estimates: Filter these first (they have line items mixed in)
        self.filtered_bills = self._filter_bills(project_names, cutoff_date)
        self.filtered_invoices = self._filter_invoices(project_names, cutoff_date)
        self.filtered_estimates = self._filter_estimates(project_names, cutoff_date)

        # Collect all organisations referenced by filtered data
        referenced_orgs = set()

        # From Projects
        for project in self.filtered_projects:
            org = project.get('Client Organisation')
            if org:
                referenced_orgs.add(org)

        # From Bills
        for bill in self.filtered_bills:
            org = bill.get('Contact Organisation')
            if org:
                referenced_orgs.add(org)

        # From Invoices
        for invoice in self.filtered_invoices:
            org = invoice.get('Contact Organisation')
            if org:
                referenced_orgs.add(org)

        # Contacts: Keep only those from referenced organisations
        self.filtered_contacts = [
            c for c in self.loader.sheets_data.get('Contacts', [])
            if c.get('Organisation') in referenced_orgs
        ]

        # Tasks: Keep those linked to kept projects
        self.filtered_tasks = [
            t for t in self.loader.sheets_data.get('Tasks', [])
            if t.get('Project') in project_names
        ]

        # Timeslips: Keep those linked to kept projects
        self.filtered_timeslips = [
            t for t in self.loader.sheets_data.get('Timeslips', [])
            if t.get('Project') in project_names
        ]

        # Price List Items: Keep all
        self.filtered_price_list = self.loader.sheets_data.get('Price List Items', [])

        if self.verbose:
            print(f"  Referenced organisations: {len(referenced_orgs)}")
            print(f"  Filtered contacts: {len(self.filtered_contacts)}")
            print(f"  Filtered tasks: {len(self.filtered_tasks)}")
            print(f"  Filtered timeslips: {len(self.filtered_timeslips)}")
            print(f"  Filtered bills: {len(self.filtered_bills)}")
            print(f"  Filtered invoices: {len(self.filtered_invoices)}")
            print(f"  Filtered estimates: {len(self.filtered_estimates)}")

    def _filter_invoices(self, project_names: set, cutoff_date: datetime) -> List[Dict]:
        """Filter invoice records (header + line items)."""
        invoices = []
        current_invoice = None

        for row in self.loader.sheets_data.get('Invoices', []):
            # Check if this is a header row (has Contact Organisation)
            if row.get('Contact Organisation'):
                # Save previous invoice if exists
                if current_invoice:
                    # Check if we should keep this invoice
                    projects = current_invoice.get('Projects', '')
                    if projects and any(p.strip() in project_names for p in projects.split(',')):
                        invoices.append(current_invoice)
                    elif current_invoice.get('Date') and isinstance(current_invoice.get('Date'), datetime) and current_invoice.get('Date') >= cutoff_date:
                        invoices.append(current_invoice)

                # Start new invoice
                current_invoice = row.copy()
                current_invoice['_line_items'] = []

            # Check if this is a line item row (has Item Type but no Contact Organisation)
            elif row.get('Item Type') and current_invoice:
                current_invoice['_line_items'].append(row)

        # Don't forget the last invoice
        if current_invoice:
            projects = current_invoice.get('Projects', '')
            if projects and any(p.strip() in project_names for p in projects.split(',')):
                invoices.append(current_invoice)
            elif current_invoice.get('Date') and isinstance(current_invoice.get('Date'), datetime) and current_invoice.get('Date') >= cutoff_date:
                invoices.append(current_invoice)

        return invoices

    def _filter_estimates(self, project_names: set, cutoff_date: datetime) -> List[Dict]:
        """Filter estimate records (header + line items)."""
        estimates = []
        current_estimate = None

        for row in self.loader.sheets_data.get('Estimates', []):
            # Check if this is a header row (has Reference - the estimate number)
            if row.get('Reference'):
                # Save previous estimate if exists
                if current_estimate:
                    project = current_estimate.get('Project', '')
                    if project and project in project_names:
                        estimates.append(current_estimate)
                    elif current_estimate.get('Date') and isinstance(current_estimate.get('Date'), datetime) and current_estimate.get('Date') >= cutoff_date:
                        estimates.append(current_estimate)

                # Start new estimate
                current_estimate = row.copy()
                current_estimate['_line_items'] = []

            # Check if this is a line item row (has Item Type but no Reference)
            elif row.get('Item Type') and current_estimate:
                current_estimate['_line_items'].append(row)

        # Don't forget the last estimate
        if current_estimate:
            project = current_estimate.get('Project', '')
            if project and project in project_names:
                estimates.append(current_estimate)
            elif current_estimate.get('Date') and isinstance(current_estimate.get('Date'), datetime) and current_estimate.get('Date') >= cutoff_date:
                estimates.append(current_estimate)

        return estimates

    def _filter_bills(self, project_names: set, cutoff_date: datetime) -> List[Dict]:
        """Filter bill records (header + line items)."""
        bills = []
        current_bill = None

        for row in self.loader.sheets_data.get('Bills', []):
            # Check if this is a header row (has Contact Organisation)
            if row.get('Contact Organisation'):
                # Save previous bill if exists
                if current_bill:
                    project = current_bill.get('Project', '')
                    if project in project_names:
                        bills.append(current_bill)
                    elif current_bill.get('Date') and isinstance(current_bill.get('Date'), datetime) and current_bill.get('Date') >= cutoff_date:
                        bills.append(current_bill)

                # Start new bill
                current_bill = row.copy()
                current_bill['_line_items'] = []

            # Check if this is a line item row (has Item Type but no Contact Organisation)
            elif row.get('Item Type') and current_bill:
                current_bill['_line_items'].append(row)

        # Don't forget the last bill
        if current_bill:
            project = current_bill.get('Project', '')
            if project in project_names:
                bills.append(current_bill)
            elif current_bill.get('Date') and isinstance(current_bill.get('Date'), datetime) and current_bill.get('Date') >= cutoff_date:
                bills.append(current_bill)

        return bills

    def _build_all_objects(self):
        """Build all Django objects in dependency order."""
        self._build_businesses()
        self._build_contacts()
        self._build_jobs_and_workorders()
        self._build_purchase_orders_and_bills()
        self._build_tasks()
        self._build_estimates()
        self._build_invoices()
        self._build_bleps()
        self._build_price_list_items()

    def _build_businesses(self):
        """Create Business objects from Contacts sheet."""
        if self.verbose:
            print("  Building businesses...")

        seen_orgs = set()
        for contact_row in self.filtered_contacts:
            org = contact_row.get('Organisation')
            if org and org not in seen_orgs:
                seen_orgs.add(org)
                pk = self.get_next_pk('contacts.business')
                self.business_map[org] = pk

                self.add_fixture('contacts.business', pk, {
                    'business_name': org,
                    'business_address': contact_row.get('Address 1', '') or '',
                    'business_number': contact_row.get('Phone Number', '') or '',
                    'tax_exemption_number': contact_row.get('Contact VAT Number', '') or '',
                    'our_reference_code': '',  # Will be auto-generated by model
                    'tax_cloud': '',
                    'terms': None,
                })

        if self.verbose:
            print(f"    Created {len(self.business_map)} businesses")

    def _build_contacts(self):
        """Create Contact objects from Contacts sheet."""
        if self.verbose:
            print("  Building contacts...")

        for contact_row in self.filtered_contacts:
            first_name = contact_row.get('First Name', '').strip() if contact_row.get('First Name') else ''
            last_name = contact_row.get('Last Name', '').strip() if contact_row.get('Last Name') else ''
            org = contact_row.get('Organisation')

            # Handle missing names
            if org and not first_name and not last_name:
                first_name = '(unknown)'
                last_name = '(unknown)'

            # Skip if no name at all
            if not first_name and not last_name:
                continue

            pk = self.get_next_pk('contacts.contact')
            full_name = f"{first_name} {last_name}".strip()

            # Map contact
            if org:
                self.contact_map[(org, full_name)] = pk

            # Get business FK
            business_fk = self.business_map.get(org) if org else None

            # Truncate phone numbers to 20 characters (max field length)
            work_phone = contact_row.get('Phone Number', '') or ''
            mobile_phone = contact_row.get('Mobile Phone Number', '') or ''
            work_number = str(work_phone)[:20] if work_phone else ''
            mobile_number = str(mobile_phone)[:20] if mobile_phone else ''

            self.add_fixture('contacts.contact', pk, {
                'name': full_name,
                'email': contact_row.get('Email', '') or '',
                'work_number': work_number,
                'mobile_number': mobile_number,
                'home_number': '',
                'addr1': contact_row.get('Address 1', '') or '',
                'addr2': contact_row.get('Address 2', '') or '',
                'addr3': contact_row.get('Address 3', '') or '',
                'city': contact_row.get('Town', '') or '',
                'municipality': contact_row.get('Region', '') or '',
                'postal_code': contact_row.get('Postcode', '') or '',
                'country_code': 'US',  # All contacts are in USA
                'business': business_fk,
            })

        if self.verbose:
            print(f"    Created {len(self.contact_map)} contacts")

    def _resolve_contact(
        self,
        org: str,
        contact_name: str,
        sheet: str,
        row: int,
        context: str
    ) -> Optional[int]:
        """
        Resolve contact reference, handling mismatches interactively.
        Returns contact PK or None.
        """
        if not org:
            return None

        # Try exact match first
        if (org, contact_name) in self.contact_map:
            return self.contact_map[(org, contact_name)]

        # Check if business exists
        if org not in self.business_map:
            if self.verbose:
                print(f"    Warning: Business '{org}' not found (referenced in {sheet} row {row})")
            return None

        # Find any contact for this business
        existing_contact = None
        existing_name = None
        existing_email = None

        for (map_org, map_name), contact_pk in self.contact_map.items():
            if map_org == org:
                existing_contact = contact_pk
                existing_name = map_name
                # Find email from fixture data
                for fixture in self.fixture_data:
                    if fixture['model'] == 'contacts.contact' and fixture['pk'] == contact_pk:
                        existing_email = fixture['fields'].get('email', '')
                        break
                break

        if not existing_contact:
            if self.verbose:
                print(f"    Warning: No contact found for business '{org}' (referenced in {sheet} row {row})")
            return None

        # Mismatch detected - prompt user
        decision = self.contact_handler.prompt_for_decision(
            business=org,
            existing_name=existing_name,
            existing_email=existing_email,
            referenced_name=contact_name,
            sheet=sheet,
            row=row,
            context=context
        )

        if decision == 'update':
            # Update existing contact's name
            self.contact_updates[existing_contact] = contact_name
            # Update in fixture data
            for fixture in self.fixture_data:
                if fixture['model'] == 'contacts.contact' and fixture['pk'] == existing_contact:
                    fixture['fields']['name'] = contact_name
                    break
            # Update map
            self.contact_map[(org, contact_name)] = existing_contact
            return existing_contact

        elif decision == 'create':
            # Create new contact
            pk = self.get_next_pk('contacts.contact')

            # Invent email/phone based on existing contact
            base_email = existing_email
            new_email = base_email  # Could modify this to make unique

            self.add_fixture('contacts.contact', pk, {
                'name': contact_name,
                'email': new_email,
                'work_number': '',
                'mobile_number': '',
                'home_number': '',
                'addr1': '',
                'addr2': '',
                'addr3': '',
                'city': '',
                'municipality': '',
                'postal_code': '',
                'country_code': 'US',  # All contacts are in USA
                'business': self.business_map[org],
            })

            self.contact_map[(org, contact_name)] = pk
            return pk

        else:  # 'map'
            # Use existing contact as-is
            self.contact_map[(org, contact_name)] = existing_contact
            return existing_contact

    def _get_v1_estimate_dates(self) -> Dict[str, datetime]:
        """
        Pre-scan estimates to find V1 estimate dates for each project.
        Returns a map of project_name -> date of V1 estimate.
        """
        v1_dates = {}

        for estimate in self.filtered_estimates:
            project_name = estimate.get('Project')
            if not project_name:
                continue

            # Parse revision from reference
            reference = estimate.get('Reference', '') or ''
            base_ref, revision = self._parse_revision_suffix(reference)

            # Only interested in V1 estimates
            if revision == 1:
                est_date = estimate.get('Date')
                if est_date and isinstance(est_date, datetime):
                    # If we already have a V1 date for this project, use the earliest one
                    if project_name not in v1_dates or est_date < v1_dates[project_name]:
                        v1_dates[project_name] = est_date

        return v1_dates

    def _has_estimates(self, project_name: str) -> bool:
        """Check if a project has any estimates."""
        for estimate in self.filtered_estimates:
            if estimate.get('Project') == project_name:
                return True
        return False

    def _build_jobs_and_workorders(self):
        """Create Job and WorkOrder objects from Projects sheet."""
        if self.verbose:
            print("  Building jobs and work orders...")

        status_map = {
            'Completed': 'completed',
            'Active': 'approved',
            'Cancelled': 'cancelled',
        }

        # Pre-scan estimates to get V1 dates for start_date calculation
        v1_estimate_dates = self._get_v1_estimate_dates()

        # Track job number counters per year
        job_counters = {}

        for project in self.filtered_projects:
            project_name = project.get('Name')
            if not project_name:
                continue

            # Create Job
            job_pk = self.get_next_pk('jobs.job')

            # Resolve contact - required for Job model
            client_org = project.get('Client Organisation')
            client_name = project.get('Client Name')
            contact_fk = None

            # Try normal resolution first (org + name)
            if client_org and client_name:
                contact_fk = self._resolve_contact(
                    org=client_org,
                    contact_name=client_name,
                    sheet='Projects',
                    row=project.get('_row', 0),
                    context=project_name
                )

            # If no org or normal resolution failed, try to find contact by name only
            if not contact_fk and client_name:
                # Search through contact_map for matching name (regardless of org)
                for (org, name), pk in self.contact_map.items():
                    if name == client_name:
                        contact_fk = pk
                        if self.verbose:
                            print(f"    Found contact '{client_name}' by name match (org='{org}')")
                        break

            # Skip this job if we still couldn't resolve a contact (Job model requires contact)
            if not contact_fk:
                if self.verbose:
                    print(f"    Skipping job '{project_name}' - no valid contact (org='{client_org}', name='{client_name}')")
                continue

            business_fk = self.business_map.get(client_org) if client_org else None

            job_status = status_map.get(project.get('Status'), 'active')

            # Calculate dates based on new rules:
            # created_date: Created Date from spreadsheet
            created_date = self._format_date(project.get('Created Date'))

            # start_date: If project has explicit "Starts On", use it;
            #             otherwise if approved, use V1 estimate date;
            #             otherwise if completed and no estimates, use created_date
            start_date = self._format_date(project.get('Starts On'))
            if not start_date and job_status == 'approved':
                # Use V1 estimate date if available
                v1_date = v1_estimate_dates.get(project_name)
                if v1_date:
                    start_date = self._format_date(v1_date)

            if not start_date and job_status == 'completed':
                # If no estimates exist for this project, use created_date
                if not self._has_estimates(project_name):
                    start_date = created_date

            # due_date: If project has explicit "Ends On", use it; otherwise leave blank
            due_date = self._format_date(project.get('Ends On'))

            # completed_date: If approved (Active), leave blank; otherwise use Updated Date
            if job_status == 'approved':
                completed_date = None
            else:
                completed_date = self._format_date(project.get('Updated Date'))

            # Generate job number in format J{year}-{counter:04d}
            created_dt = project.get('Created Date')
            if isinstance(created_dt, datetime):
                year = created_dt.year
            else:
                year = 2025  # Default year if no date

            # Increment counter for this year
            if year not in job_counters:
                job_counters[year] = 1
            else:
                job_counters[year] += 1

            job_number = f"J{year}-{job_counters[year]:04d}"

            self.add_fixture('jobs.job', job_pk, {
                'name': project_name,
                'job_number': job_number,
                'contact': contact_fk,
                'start_date': start_date,
                'due_date': due_date,
                'created_date': created_date,
                'customer_po_number': project.get('Contract PO Reference', '') or '',
                'status': job_status,
                'description': project.get('Notes', '') or '',
                'completed_date': completed_date,
            })

            # Create WorkOrder if not cancelled
            workorder_pk = None
            if project.get('Status') != 'CANCELLED':
                workorder_pk = self.get_next_pk('jobs.workorder')

                wo_status = 'complete' if job_status == 'completed' else 'draft'

                self.add_fixture('jobs.workorder', workorder_pk, {
                    'job': job_pk,
                    'status': wo_status,
                    'template': None,
                })

            self.job_map[project_name] = (job_pk, workorder_pk)

        if self.verbose:
            jobs_count = len(self.job_map)
            wo_count = sum(1 for _, wo_pk in self.job_map.values() if wo_pk is not None)
            print(f"    Created {jobs_count} jobs and {wo_count} work orders")

    def _build_purchase_orders_and_bills(self):
        """Create PurchaseOrder and Bill objects from Bills sheet."""
        if self.verbose:
            print("  Building purchase orders and bills...")

        # Track bill number counters per year
        bill_counters = {}
        po_counters = {}

        for bill in self.filtered_bills:
            line_items = bill.get('_line_items', [])
            self._save_bill_and_po(bill, line_items, bill_counters, po_counters)

        if self.verbose:
            print(f"    Created {self.pk_counters.get('purchasing.purchaseorder', 1) - 1} purchase orders")
            print(f"    Created {self.pk_counters.get('purchasing.bill', 1) - 1} bills")

    def _save_bill_and_po(self, bill_row: Dict, line_items: List[Dict], bill_counters: Dict, po_counters: Dict):
        """Save a Bill and its associated PurchaseOrder and line items."""
        # Create PurchaseOrder
        po_pk = self.get_next_pk('purchasing.purchaseorder')

        # Resolve contact
        contact_org = bill_row.get('Contact Organisation')
        contact_name = bill_row.get('Contact Name')
        contact_fk = None

        if contact_org and contact_name:
            contact_fk = self._resolve_contact(
                org=contact_org,
                contact_name=contact_name,
                sheet='Bills',
                row=bill_row.get('_row', 0),
                context=bill_row.get('Reference', 'Unknown')
            )

        business_fk = self.business_map.get(contact_org) if contact_org else None

        # Get job FK
        project_name = bill_row.get('Project')
        job_fk = None
        if project_name and project_name in self.job_map:
            job_fk, _ = self.job_map[project_name]

        # Dates
        created_date = self._format_date(bill_row.get('Date'))

        # Generate PO number in format PO{year}-{counter:04d}
        created_dt = bill_row.get('Date')
        if isinstance(created_dt, datetime):
            year = created_dt.year
        else:
            year = 2025  # Default year if no date

        # Increment PO counter for this year
        if year not in po_counters:
            po_counters[year] = 1
        else:
            po_counters[year] += 1

        po_number = f"PO{year}-{po_counters[year]:04d}"

        self.add_fixture('purchasing.purchaseorder', po_pk, {
            'po_number': po_number,
            'business': business_fk,
            'contact': contact_fk,
            'job': job_fk,
            'status': 'issued',  # All imported POs are issued
            'created_date': created_date,
            'requested_date': None,
            'issued_date': created_date,
            'received_date': None,
            'cancel_date': None,
        })

        # Create Bill
        bill_pk = self.get_next_pk('purchasing.bill')

        due_date = self._format_date(bill_row.get('Due Date'))
        reference = bill_row.get('Reference', '')

        # Generate Bill number in format B{year}-{counter:04d}
        if year not in bill_counters:
            bill_counters[year] = 1
        else:
            bill_counters[year] += 1

        bill_number = f"B{year}-{bill_counters[year]:04d}"

        self.add_fixture('purchasing.bill', bill_pk, {
            'bill_number': bill_number,
            'purchase_order': po_pk,
            'business': business_fk,
            'contact': contact_fk,
            'vendor_invoice_number': reference or '',
            'status': 'draft',
            'created_date': created_date,
            'due_date': due_date,
            'received_date': None,
            'paid_date': None,
            'cancelled_date': None,
        })

        # Create line items for both PO and Bill with sequential line numbers starting at 1
        line_number = 1
        for item_row in line_items:
            # PO Line Item
            po_item_pk = self.get_next_pk('purchasing.purchaseorderlineitem')

            qty = self._parse_decimal(item_row.get('Quantity', 1))
            price = self._parse_decimal(item_row.get('Net Value', 0))

            self.add_fixture('purchasing.purchaseorderlineitem', po_item_pk, {
                'purchase_order': po_pk,
                'task': None,
                'price_list_item': None,
                'line_number': line_number,
                'qty': str(qty),
                'units': item_row.get('Item Type', '-no unit-') or '-no unit-',
                'description': item_row.get('Description', '') or '',
                'price': str(price),
            })

            # Bill Line Item
            bill_item_pk = self.get_next_pk('purchasing.billlineitem')

            self.add_fixture('purchasing.billlineitem', bill_item_pk, {
                'bill': bill_pk,
                'task': None,
                'price_list_item': None,
                'line_number': line_number,
                'qty': str(qty),
                'units': item_row.get('Item Type', '-no unit-') or '-no unit-',
                'description': item_row.get('Description', '') or '',
                'price': str(price),
            })

            line_number += 1

    def _build_tasks(self):
        """Create Task objects from Tasks sheet."""
        if self.verbose:
            print("  Building tasks...")

        # Group tasks by project/work order
        tasks_by_wo = {}

        for task_row in self.filtered_tasks:
            project_name = task_row.get('Project')
            if not project_name or project_name not in self.job_map:
                continue

            _, workorder_pk = self.job_map[project_name]
            if not workorder_pk:
                continue  # Skip if no work order (cancelled project)

            if workorder_pk not in tasks_by_wo:
                tasks_by_wo[workorder_pk] = []

            tasks_by_wo[workorder_pk].append(task_row)

        # Create tasks with line numbers
        for workorder_pk, tasks in tasks_by_wo.items():
            for line_num, task_row in enumerate(tasks, start=1):
                task_pk = self.get_next_pk('jobs.task')

                task_name = task_row.get('Name', '')
                project_name = task_row.get('Project', '')

                # Store in task map for Blep lookup
                self.task_map[(project_name, task_name)] = task_pk

                rate = self._parse_decimal(task_row.get('Billing Rate', 0))

                self.add_fixture('jobs.task', task_pk, {
                    'parent_task': None,
                    'assignee': None,
                    'work_order': workorder_pk,
                    'est_worksheet': None,
                    'name': task_name,
                    'line_number': line_num,
                    'units': 'hours',  # Default
                    'rate': str(rate),
                    'est_qty': '0',
                    'template': None,
                })

        if self.verbose:
            print(f"    Created {len(self.task_map)} tasks")

    def _create_additional_job(self, original_project_name: str, estimate_base_ref: str, part_number: int,
                                job_counters: Dict, v1_estimate_dates: Dict) -> Tuple[int, int]:
        """Create an additional job for a multi-estimate project."""
        # Find the original project data
        original_project = None
        for project in self.filtered_projects:
            if project.get('Name') == original_project_name:
                original_project = project
                break

        if not original_project:
            return None, None

        # Get original job info
        original_job_fk, _ = self.job_map.get(original_project_name, (None, None))

        # Create new job
        job_pk = self.get_next_pk('jobs.job')

        # Get contact info (same as original)
        client_org = original_project.get('Client Organisation')
        client_name = original_project.get('Client Name')
        contact_fk = None

        if client_org and client_name:
            contact_fk = self._resolve_contact(
                org=client_org,
                contact_name=client_name,
                sheet='Projects',
                row=original_project.get('_row', 0),
                context=original_project_name
            )

        if not contact_fk and client_name:
            for (org, name), pk in self.contact_map.items():
                if name == client_name:
                    contact_fk = pk
                    break

        if not contact_fk:
            return None, None

        business_fk = self.business_map.get(client_org) if client_org else None

        # Generate job number
        created_dt = original_project.get('Created Date')
        if isinstance(created_dt, datetime):
            year = created_dt.year
        else:
            year = 2025

        if year not in job_counters:
            job_counters[year] = 1
        else:
            job_counters[year] += 1

        job_number = f"J{year}-{job_counters[year]:04d}"

        # Create modified job name (max 50 chars for Job.name field)
        # Format: "Original Name - Est XXXX"
        suffix = f" - Est {estimate_base_ref}"
        max_base_len = 50 - len(suffix)

        if len(original_project_name) <= max_base_len:
            new_job_name = original_project_name + suffix
        else:
            # Truncate original name to fit
            new_job_name = original_project_name[:max_base_len] + suffix

        # Get dates and status
        job_status = original_project.get('Status', 'Active')
        status_map = {'Completed': 'completed', 'Active': 'approved', 'Cancelled': 'cancelled'}
        job_status = status_map.get(job_status, 'approved')

        # Calculate dates using same rules as main job creation
        created_date = self._format_date(original_project.get('Created Date'))

        # start_date: If project has explicit "Starts On", use it;
        #             otherwise if approved, use V1 estimate date;
        #             otherwise if completed and no estimates, use created_date
        start_date = self._format_date(original_project.get('Starts On'))
        if not start_date and job_status == 'approved':
            # Use V1 estimate date if available
            v1_date = v1_estimate_dates.get(original_project_name)
            if v1_date:
                start_date = self._format_date(v1_date)

        if not start_date and job_status == 'completed':
            # If no estimates exist for this project, use created_date
            if not self._has_estimates(original_project_name):
                start_date = created_date

        # due_date: If project has explicit "Ends On", use it; otherwise leave blank
        due_date = self._format_date(original_project.get('Ends On'))

        # completed_date: If approved (Active), leave blank; otherwise use Updated Date
        if job_status == 'approved':
            completed_date = None
        else:
            completed_date = self._format_date(original_project.get('Updated Date'))

        # Add reference to original job in description
        original_job_num = None
        for name, (_, _) in self.job_map.items():
            if name == original_project_name:
                # Find the original job number from fixtures
                for fixture in self.fixture_data:
                    if fixture['model'] == 'jobs.job' and fixture['fields'].get('name') == original_project_name:
                        original_job_num = fixture['fields'].get('job_number')
                        break
                break

        description = original_project.get('Notes', '') or ''
        if original_job_num:
            description = f"Related to Job {original_job_num}. " + description

        self.add_fixture('jobs.job', job_pk, {
            'name': new_job_name,
            'job_number': job_number,
            'contact': contact_fk,
            'start_date': start_date,
            'due_date': due_date,
            'created_date': created_date,
            'customer_po_number': original_project.get('Contract PO Reference', '') or '',
            'status': job_status,
            'description': description,
            'completed_date': completed_date,
        })

        # Create workorder
        workorder_pk = self.get_next_pk('jobs.workorder')
        wo_status = 'complete' if job_status == 'completed' else 'draft'

        self.add_fixture('jobs.workorder', workorder_pk, {
            'job': job_pk,
            'status': wo_status,
            'template': None,
        })

        return job_pk, workorder_pk

    def _build_estimates(self):
        """Create Estimate and EstimateLineItem objects."""
        if self.verbose:
            print("  Building estimates...")

        status_map = {
            'Draft': 'draft',
            'Sent': 'open',
            'Approved': 'accepted',
            'Rejected': 'rejected',
        }

        # Get V1 estimate dates for job date calculations
        v1_estimate_dates = self._get_v1_estimate_dates()

        # First, group estimates by project to detect multi-estimate situations
        estimates_by_project = {}
        for estimate in self.filtered_estimates:
            project_name = estimate.get('Project')
            if not project_name or project_name not in self.job_map:
                continue

            if project_name not in estimates_by_project:
                estimates_by_project[project_name] = []

            estimates_by_project[project_name].append(estimate)

        # Track estimate PKs by base reference and revision for parent relationships
        estimate_pk_map = {}
        estimate_count = 0
        line_item_count = 0
        additional_jobs_created = 0

        # Job counters for creating additional jobs
        job_counters = {}

        # Initialize counters from existing jobs
        for fixture in self.fixture_data:
            if fixture['model'] == 'jobs.job':
                job_num = fixture['fields'].get('job_number', '')
                if job_num.startswith('J'):
                    try:
                        parts = job_num.split('-')
                        if len(parts) == 2:
                            year = int(parts[0][1:])
                            counter = int(parts[1])
                            if year not in job_counters or counter > job_counters[year]:
                                job_counters[year] = counter
                    except:
                        pass

        # Process each project's estimates
        for project_name, project_estimates in estimates_by_project.items():
            # Group by base reference number within this project
            estimates_by_base_ref = {}

            for estimate in project_estimates:
                reference = estimate.get('Reference', '') or ''
                base_ref, revision = self._parse_revision_suffix(reference)

                if base_ref not in estimates_by_base_ref:
                    estimates_by_base_ref[base_ref] = []

                estimates_by_base_ref[base_ref].append({
                    'estimate': estimate,
                    'base_ref': base_ref,
                    'revision': revision,
                    'reference': reference,
                })

            # Sort estimates by date to determine which is "first"
            base_refs_sorted = sorted(
                estimates_by_base_ref.keys(),
                key=lambda br: min(
                    e['estimate'].get('Date') or datetime(1900, 1, 1)
                    for e in estimates_by_base_ref[br]
                )
            )

            # Track jobs for this project's estimates
            jobs_for_estimates = {}

            # First base ref uses the original job
            first_base_ref = base_refs_sorted[0]
            original_job_fk, original_wo_fk = self.job_map[project_name]
            jobs_for_estimates[first_base_ref] = (original_job_fk, original_wo_fk)

            # Additional base refs need new jobs
            for i, base_ref in enumerate(base_refs_sorted[1:], start=2):
                new_job_fk, new_wo_fk = self._create_additional_job(
                    project_name, base_ref, i, job_counters, v1_estimate_dates
                )
                if new_job_fk:
                    jobs_for_estimates[base_ref] = (new_job_fk, new_wo_fk)
                    additional_jobs_created += 1
                    if self.verbose:
                        print(f"    Created additional job for {project_name} / estimate {base_ref}")

            # Now create estimates, linking to appropriate jobs
            for base_ref in base_refs_sorted:
                if base_ref not in jobs_for_estimates:
                    continue

                job_fk, _ = jobs_for_estimates[base_ref]
                revisions = estimates_by_base_ref[base_ref]

                # Sort revisions
                revisions.sort(key=lambda x: x['revision'])
                max_revision = max(r['revision'] for r in revisions)

                for rev_data in revisions:
                    estimate = rev_data['estimate']
                    revision = rev_data['revision']

                    # Create Estimate
                    estimate_pk = self.get_next_pk('jobs.estimate')
                    estimate_count += 1

                    # Store PK for parent relationship
                    estimate_pk_map[(base_ref, revision)] = estimate_pk

                    # Determine parent FK
                    parent_fk = None
                    if revision > 1:
                        parent_fk = estimate_pk_map.get((base_ref, revision - 1))

                    # Determine status
                    status = status_map.get(estimate.get('Status'), 'draft')
                    if revision < max_revision:
                        status = 'superseded'

                    created_date = self._format_date(estimate.get('Date'))

                    self.add_fixture('jobs.estimate', estimate_pk, {
                        'job': job_fk,
                        'estimate_number': base_ref,
                        'version': revision,
                        'status': status,
                        'parent': parent_fk,
                        'created_date': created_date,
                        'sent_date': created_date if status in ['open', 'accepted', 'rejected'] else None,
                        'closed_date': None,
                        'expiration_date': None,
                    })

                    # Create line items with sequential line numbers starting at 1
                    line_number = 1
                    for item_row in estimate.get('_line_items', []):
                        line_item_pk = self.get_next_pk('jobs.estimatelineitem')
                        line_item_count += 1

                        qty = self._parse_decimal(item_row.get('Quantity', 1))
                        price = self._parse_decimal(item_row.get('Price', 0))

                        self.add_fixture('jobs.estimatelineitem', line_item_pk, {
                            'estimate': estimate_pk,
                            'task': None,
                            'price_list_item': None,
                            'line_number': line_number,
                            'qty': str(qty),
                            'units': '',
                            'description': item_row.get('Description', '') or '',
                            'price': str(price),
                        })

                        line_number += 1

        if self.verbose:
            print(f"    Created {estimate_count} estimates with {line_item_count} line items")
            if additional_jobs_created > 0:
                print(f"    Created {additional_jobs_created} additional jobs for multi-estimate projects")

    def _build_invoices(self):
        """Create Invoice and InvoiceLineItem objects."""
        if self.verbose:
            print("  Building invoices...")

        status_map = {
            'Draft': 'draft',
            'Sent': 'open',
            'Cancelled': 'cancelled',
        }

        invoice_count = 0
        line_item_count = 0

        for invoice in self.filtered_invoices:
            # Get job FK from Projects field
            projects_str = invoice.get('Projects', '')
            job_fk = None

            if projects_str:
                # Try first project name in the comma-separated list
                for project_name in projects_str.split(','):
                    project_name = project_name.strip()
                    if project_name in self.job_map:
                        job_fk, _ = self.job_map[project_name]
                        break

            if not job_fk:
                continue

            # Get business FK
            contact_org = invoice.get('Contact Organisation')
            business_fk = self.business_map.get(contact_org) if contact_org else None

            # Create Invoice
            invoice_pk = self.get_next_pk('invoicing.invoice')
            invoice_count += 1

            status = status_map.get(invoice.get('Status'), 'draft')
            created_date = self._format_date(invoice.get('Date'))
            closed_date = self._format_date(invoice.get('Paid Date')) if status == 'paid' else None

            self.add_fixture('invoicing.invoice', invoice_pk, {
                'job': job_fk,
                'invoice_number': invoice.get('Reference', '') or '',
                'status': status,
                'created_date': created_date,
                'sent_date': created_date if status in ['open', 'paid'] else None,
                'closed_date': closed_date,
            })

            # Create line items with sequential line numbers starting at 1
            line_number = 1
            for item_row in invoice.get('_line_items', []):
                line_item_pk = self.get_next_pk('invoicing.invoicelineitem')
                line_item_count += 1

                qty = self._parse_decimal(item_row.get('Quantity', 1))
                price = self._parse_decimal(item_row.get('Price', 0))

                self.add_fixture('invoicing.invoicelineitem', line_item_pk, {
                    'invoice': invoice_pk,
                    'task': None,
                    'price_list_item': None,
                    'line_number': line_number,
                    'qty': str(qty),
                    'units': '',
                    'description': item_row.get('Description', '') or '',
                    'price': str(price),
                })

                line_number += 1

        if self.verbose:
            print(f"    Created {invoice_count} invoices with {line_item_count} line items")

    def _build_bleps(self):
        """Create Blep objects from Timeslips sheet."""
        import random

        if self.verbose:
            print("  Building bleps (time tracking)...")

        for timeslip in self.filtered_timeslips:
            # Find task
            project_name = timeslip.get('Project')
            task_name = timeslip.get('Task')

            if not project_name or not task_name:
                continue

            task_pk = self.task_map.get((project_name, task_name))
            if not task_pk:
                if self.verbose:
                    print(f"    Warning: Task '{task_name}' not found for project '{project_name}'")
                continue

            # Calculate start and end times
            date = timeslip.get('Date')
            hours = timeslip.get('Hours', 0)

            if not date or not isinstance(date, datetime):
                continue

            # Start time: 9:00 AM on the date
            start_time = datetime.combine(date.date(), datetime.min.time().replace(hour=9))
            # End time: start + hours
            end_time = start_time + timedelta(hours=float(hours) if hours else 0)

            blep_pk = self.get_next_pk('jobs.blep')

            # Randomly assign user to either PK 2 or 3
            user_pk = random.choice([2, 3])

            self.add_fixture('jobs.blep', blep_pk, {
                'user': user_pk,
                'task': task_pk,
                'start_time': start_time.isoformat(),
                'end_time': end_time.isoformat(),
            })

        if self.verbose:
            blep_count = self.pk_counters.get('jobs.blep', 1) - 1
            print(f"    Created {blep_count} bleps")

    def _build_price_list_items(self):
        """Create PriceListItem objects."""
        if self.verbose:
            print("  Building price list items...")

        for item in self.filtered_price_list:
            pk = self.get_next_pk('invoicing.pricelistitem')

            qty = self._parse_decimal(item.get('Quantity', 1))
            price = self._parse_decimal(item.get('Price', 0))

            self.add_fixture('invoicing.pricelistitem', pk, {
                'code': item.get('Code', '') or '',
                'units': item.get('Type', '') or '',
                'description': item.get('Description', '') or '',
                'purchase_price': str(price),
                'selling_price': str(price),
                'qty_on_hand': str(qty),
                'qty_sold': '0',
                'qty_wasted': '0',
                'is_active': True,
            })

        if self.verbose:
            print(f"    Created {len(self.filtered_price_list)} price list items")

    def _format_date(self, value) -> Optional[str]:
        """Format date value to ISO string."""
        if not value:
            return None

        if isinstance(value, datetime):
            return value.date().isoformat()

        return None

    def _parse_decimal(self, value) -> Decimal:
        """Parse decimal value."""
        if value is None:
            return Decimal('0')

        try:
            return Decimal(str(value))
        except:
            return Decimal('0')

    def _parse_revision_suffix(self, reference: str) -> tuple:
        """
        Parse revision suffix from reference number.

        Examples:
            'EST123' -> ('EST123', 1)
            'EST123-r2' -> ('EST123', 2)
            'EST123-r3' -> ('EST123', 3)
            'EST123-rev2' -> ('EST123', 2)

        Returns:
            (base_reference, revision_number)
        """
        import re

        if not reference:
            return ('', 1)

        # Match patterns like -r2, -r3, -rev2, -rev3
        match = re.search(r'-r(?:ev)?(\d+)$', reference, re.IGNORECASE)

        if match:
            revision = int(match.group(1))
            base_ref = reference[:match.start()]
            return (base_ref, revision)

        # No revision suffix found
        return (reference, 1)

    def _print_summary(self):
        """Print summary of objects to be created."""
        print("=" * 70)
        print("Summary")
        print("=" * 70)

        model_counts = {}
        for fixture in self.fixture_data:
            model = fixture['model']
            model_counts[model] = model_counts.get(model, 0) + 1

        print("\nObjects to be created:")
        for model in sorted(model_counts.keys()):
            print(f"  {model:40} {model_counts[model]:5}")

        print(f"\n  {'TOTAL':40} {len(self.fixture_data):5}")

    def _write_json(self):
        """Write fixture data to JSON file."""
        with open(self.output_path, 'w') as f:
            json.dump(self.fixture_data, f, indent=2, default=str)


def main():
    parser = argparse.ArgumentParser(
        description='Convert Neal\'s CNC Excel export to Django fixture JSON',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s nealsdata/company-export.xlsx
  %(prog)s nealsdata/company-export.xlsx --output my_data.json
  %(prog)s nealsdata/company-export.xlsx --non-interactive
  %(prog)s nealsdata/company-export.xlsx --dry-run --verbose
        """
    )

    parser.add_argument(
        'excel_file',
        help='Path to Excel file to convert'
    )

    parser.add_argument(
        '--output',
        default='neals_data.json',
        help='Output JSON file path (default: neals_data.json)'
    )

    parser.add_argument(
        '--base-fixture',
        default='fixtures/job_data/01_base.json',
        help='Base fixture file to load first (default: fixtures/job_data/01_base.json)'
    )

    parser.add_argument(
        '--non-interactive',
        action='store_true',
        help='Skip interactive prompts, auto-map all contact mismatches'
    )

    parser.add_argument(
        '--dry-run',
        action='store_true',
        help='Show statistics without generating JSON file'
    )

    parser.add_argument(
        '--verbose',
        action='store_true',
        help='Show detailed progress information'
    )

    args = parser.parse_args()

    # Check if file exists
    if not Path(args.excel_file).exists():
        print(f"Error: File not found: {args.excel_file}")
        sys.exit(1)

    # Run converter
    converter = NealsDataConverter(
        excel_path=args.excel_file,
        output_path=args.output,
        base_fixture_path=args.base_fixture,
        interactive=not args.non_interactive,
        dry_run=args.dry_run,
        verbose=args.verbose
    )

    try:
        converter.convert()
    except KeyboardInterrupt:
        print("\n\nConversion cancelled by user.")
        sys.exit(1)
    except Exception as e:
        print(f"\n\nError during conversion: {e}")
        if args.verbose:
            import traceback
            traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
